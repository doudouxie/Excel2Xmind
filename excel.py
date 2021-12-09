import openpyxl

wb= openpyxl.load_workbook('test.xlsx')
ws = wb.active

data = [1,2,3,4,5]

# def liebiao(list,dict={}):
#     while len(list)>1:
#          if len(list)==2:
#              dict_={}
#              dict_[list[0]]=list[1]
#              return dict_
#          key = list.pop(0)
#          dict[key]=liebiao(list)
#          print(dict)
#
#
# print(liebiao(data))

def hehe(a):
    d={}
    if len(a)>2:
        d[a[0]]=hehe(a[1:])
    else:
        d[a[0]]=a[1]
    return d

data = []
for row in range(2,ws.max_row+1):
    if ws[row][4].value is None:
        ws[row][4].value = 'Null'
    if ws[row][5].value is None:
        ws[row][5].value = 'Null'
    if ws[row][3].value is not None:
        case = "/*"+ws[row][6].value+"*/"+ws[row][2].value+"/*"+ws[row][3].value+"*/"
    else:
        case = "/*" + ws[row][6].value + "*/" + ws[row][2].value
        #print(case)


    pre_result = ws[row][0].value+"$"+ws[row][1].value+"$"+"cases"+"$"+case
    if '\n' in ws[row][4].value:
        miaoshu = ws[row][4].value.split('\n')
        jieguo = ws[row][5].value.split('\n')
        for i in range(len(miaoshu)):
            if i<len(jieguo):
                result = (pre_result+"$"+miaoshu[i]+"$"+jieguo[i])

            else:
                result = (pre_result + "$" + miaoshu[i])
            after = result.split("$")
            # print(result)
            # print(after)
            final = hehe(after)
            data.append(final)
        #after = result.split("$")
        #print(hehe(after)
        # data.append(hehe(after))
    else:
        result= pre_result+"$"+ws[row][4].value+"$"+ws[row][5].value
        after = result.split("$")
        # print(hehe(after))
        # data.append(hehe(after))
        final = hehe(after)
        #print(final)
        data.append(final)
print(data)












